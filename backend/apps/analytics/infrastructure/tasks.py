import json
from pathlib import Path
from uuid import uuid4

from celery import shared_task
from django.conf import settings
from openpyxl import Workbook
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.legends import Legend
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing, String
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.commerce.infrastructure.models import Order

from ..application.queries import DashboardQuery, InventoryReportQuery, SalesReportQuery

CHART_COLORS = (
    colors.HexColor("#0a0a0a"),
    colors.HexColor("#4a4a4a"),
    colors.HexColor("#6a6a6a"),
    colors.HexColor("#8a8a8a"),
    colors.HexColor("#aaaaaa"),
    colors.HexColor("#cacaca"),
    colors.HexColor("#dadada"),
)

MONTH_LABELS = {
    "01": "Ene", "02": "Feb", "03": "Mar", "04": "Abr", "05": "May", "06": "Jun",
    "07": "Jul", "08": "Ago", "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dic",
}


def _month_label(month):
    return MONTH_LABELS.get(month.split("-")[-1], month)


def _build_period_label(filters):
    """Arma un texto legible con los filtros aplicados, para mostrarlo en el archivo exportado."""
    parts = []
    date_from = filters.get("date_from")
    date_to = filters.get("date_to")
    if date_from or date_to:
        parts.append(f"{date_from or 'inicio'} a {date_to or 'hoy'}")
    status = filters.get("status")
    if status:
        parts.append(f"Estado: {dict(Order.Status.choices).get(status, status)}")
    client_type = filters.get("client_type")
    if client_type:
        client_label = "Mayorista" if client_type == "WHOLESALE" else "Retail"
        parts.append(f"Tipo de cliente: {client_label}")
    return " · ".join(parts) if parts else "Todos los datos (sin filtros)"


INVENTORY_REPORTS = {
    "compras", "bajo-minimo", "produccion", "mermas", "corte",
    "inv-general", "inv-bodega", "inv-grupo", "valorizado", "movimientos",
}

INVENTORY_REPORT_TITLES = {
    "compras": "Órdenes de Compra",
    "bajo-minimo": "Artículos bajo Mínimo",
    "produccion": "Órdenes de Producción",
    "mermas": "Mermas y Sobrantes",
    "corte": "Inventario a Fecha de Corte",
    "inv-general": "Inventario General",
    "inv-bodega": "Inventario por Bodega",
    "inv-grupo": "Inventario por Grupo",
    "valorizado": "Valorización de Inventario",
    "movimientos": "Movimientos del Período",
}


def _build_inventory_period_label(filters):
    parts = []
    date_from = filters.get("date_from") or filters.get("desde")
    date_to = filters.get("date_to") or filters.get("hasta")
    if date_from or date_to:
        parts.append(f"{date_from or 'inicio'} a {date_to or 'hoy'}")
    bodega = filters.get("bodega")
    if bodega:
        parts.append(f"Bodega: {bodega}")
    grupo = filters.get("grupo")
    if grupo:
        parts.append(f"Grupo: {grupo}")
    return " · ".join(parts) if parts else "Todos los datos (sin filtros)"


def _generate_inventory_report(output_path, report_type, output_format, filters):
    query = InventoryReportQuery()
    date_from = filters.get("date_from") or filters.get("desde") or None
    date_to = filters.get("date_to") or filters.get("hasta") or None
    bodega = filters.get("bodega") or None
    grupo = filters.get("grupo") or None
    period_label = _build_inventory_period_label(filters)
    title = INVENTORY_REPORT_TITLES[report_type]

    if report_type == "compras":
        data = query.purchases(date_from=date_from, date_to=date_to, bodega=bodega, grupo=grupo)
    elif report_type == "bajo-minimo":
        data = query.low_stock(bodega=bodega, grupo=grupo)
    elif report_type == "produccion":
        data = query.production(date_from=date_from, date_to=date_to, grupo=grupo)
    elif report_type == "mermas":
        data = query.losses(date_from=date_from, date_to=date_to, bodega=bodega, grupo=grupo)
    elif report_type == "corte":
        from datetime import date as date_cls
        cutoff = date_to or date_from or date_cls.today().isoformat()
        data = query.stock_at_date(cutoff, bodega=bodega, grupo=grupo)
    elif report_type == "inv-general":
        data = query.stock_general(bodega=bodega, grupo=grupo)
    elif report_type == "inv-bodega":
        data = query.stock_by_warehouse(bodega=bodega, grupo=grupo)
    elif report_type == "inv-grupo":
        data = query.stock_by_group(bodega=bodega, grupo=grupo)
    elif report_type == "valorizado":
        data = query.valuation(bodega=bodega, grupo=grupo)
    else:  # movimientos
        data = query.movements(date_from=date_from, date_to=date_to, bodega=bodega, grupo=grupo)

    if output_format == "pdf":
        _write_inventory_report_pdf(output_path, report_type, title, data, period_label)
    else:
        _write_inventory_report_xlsx(output_path, report_type, title, data, period_label)


@shared_task
def generate_report(report_type, output_format="xlsx", filters=None):
    filters = filters or {}
    report_dir = Path(settings.MEDIA_ROOT) / "reports"
    report_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{report_type}-{uuid4().hex}.{output_format}"
    output_path = report_dir / filename

    CUSTOMER_REPORTS = {"customers", "customer_geo", "international_customers"}
    if report_type in CUSTOMER_REPORTS:
        sales_data = SalesReportQuery().execute(
            date_from=filters.get("date_from"),
            date_to=filters.get("date_to"),
            status=filters.get("status"),
            client_type=filters.get("client_type"),
        )
        period_label = _build_period_label(filters)
        if output_format == "pdf":
            _write_customer_report_pdf(output_path, report_type, sales_data, period_label)
        else:
            if report_type == "customers":
                _write_customers_xlsx(output_path, sales_data, period_label)
            elif report_type == "customer_geo":
                _write_customer_geo_xlsx(output_path, sales_data, period_label)
            else:
                _write_international_customers_xlsx(output_path, sales_data, period_label)
    elif report_type in INVENTORY_REPORTS:
        _generate_inventory_report(output_path, report_type, output_format, filters)
    else:
        data = DashboardQuery().execute()
        if output_format == "pdf":
            _write_pdf(output_path, report_type, data)
        else:
            _write_xlsx(output_path, report_type, data)

    return {
        "report_type": report_type,
        "filters": filters or {},
        "status": "generated",
        "url": f"{settings.MEDIA_URL}reports/{filename}",
    }


def _fmt_cop(value):
    try:
        return f"$ {float(value):,.0f}".replace(",", ".")
    except (TypeError, ValueError):
        return str(value)


def _xlsx_header_style():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    font = Font(bold=True, color="FFFFFF", size=10)
    fill = PatternFill("solid", fgColor="1A1A1A")
    alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return font, fill, alignment, border


def _xlsx_row_style(even):
    from openpyxl.styles import PatternFill, Alignment, Border, Side
    fill = PatternFill("solid", fgColor="F5F5F4" if even else "FFFFFF")
    alignment = Alignment(vertical="center")
    thin = Side(style="thin", color="E7E5E4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return fill, alignment, border


def _write_period_label(sheet, period_label):
    """Escribe el período exportado en la fila 1 y devuelve la fila donde debe ir el header."""
    if not period_label:
        return 1
    from openpyxl.styles import Font
    sheet.cell(row=1, column=1, value=f"Período exportado: {period_label}").font = Font(italic=True, color="6B7280")
    return 3


def _apply_header(sheet, headers, start_row=1):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    font, fill, alignment, border = _xlsx_header_style()
    sheet.row_dimensions[start_row].height = 20
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=start_row, column=col_idx, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border


def _apply_row(sheet, row_idx, values):
    fill, alignment, border = _xlsx_row_style(row_idx % 2 == 0)
    for col_idx, value in enumerate(values, 1):
        cell = sheet.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border


def _write_customers_xlsx(path, data, period_label=""):
    from openpyxl.styles import Font
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Clientes activos"

    header_row = _write_period_label(sheet, period_label)
    headers = ["#", "Nombre", "Email", "Ciudad", "Pedidos", "Ingresos", "Ticket promedio", "Última compra", "Segmento", "Modo"]
    _apply_header(sheet, headers, start_row=header_row)

    segment_map = {"VIP": "★ VIP", "Recurrente": "Recurrente", "Nuevo": "Nuevo", "Inactivo": "Inactivo"}
    mode_map = {"WHOLESALE": "Mayorista", "RETAIL": "Retail"}

    for i, c in enumerate(data.get("top_customers", []), 1):
        row_idx = header_row + i
        last_order = ""
        if c.get("last_order"):
            try:
                from datetime import datetime
                last_order = datetime.fromisoformat(c["last_order"]).strftime("%d/%m/%Y")
            except Exception:
                last_order = c["last_order"]
        values = (
            i, c["name"], c["email"], c["city"] or "—",
            c["orders"], _fmt_cop(c["revenue"]), _fmt_cop(c["avg_ticket"]),
            last_order, segment_map.get(c["segment"], c["segment"]),
            mode_map.get(c["mode"], c["mode"]),
        )
        _apply_row(sheet, row_idx, values)
        if c["segment"] == "VIP":
            sheet.cell(row=row_idx, column=9).font = Font(bold=True, color="B45309")

    _auto_width(sheet)

    # hoja de segmentación
    seg_sheet = workbook.create_sheet("Segmentación")
    seg_header_row = _write_period_label(seg_sheet, period_label)
    _apply_header(seg_sheet, ["Segmento", "Cantidad", "Porcentaje"], start_row=seg_header_row)
    for i, s in enumerate(data.get("customer_segments", []), 1):
        _apply_row(seg_sheet, seg_header_row + i, (s["segment"], s["count"], f"{s['percentage']}%"))
    _auto_width(seg_sheet)

    workbook.save(path)


def _write_customer_geo_xlsx(path, data, period_label=""):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Distribución geográfica"

    header_row = _write_period_label(sheet, period_label)
    headers = ["#", "Ciudad", "Clientes", "Pedidos", "Ingresos totales", "% de la base"]
    _apply_header(sheet, headers, start_row=header_row)

    for i, g in enumerate(data.get("customer_geo", []), 1):
        _apply_row(sheet, header_row + i, (
            i, g["city"], g["customers"], g["orders"],
            _fmt_cop(g["revenue"]), f"{g['percentage']}%",
        ))

    _auto_width(sheet)
    workbook.save(path)


def _write_international_customers_xlsx(path, data, period_label=""):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Clientes internacionales"

    header_row = _write_period_label(sheet, period_label)
    headers = ["#", "Nombre", "Email", "País(es)", "Pedidos", "Ingresos", "Modo", "Distribuidor"]
    _apply_header(sheet, headers, start_row=header_row)

    mode_map = {"WHOLESALE": "Mayorista", "RETAIL": "Retail"}
    for i, c in enumerate(data.get("international_customers", []), 1):
        _apply_row(sheet, header_row + i, (
            i, c["name"], c["email"],
            ", ".join(c["countries"]),
            c["orders"], _fmt_cop(c["revenue"]),
            mode_map.get(c["mode"], c["mode"]),
            "Sí" if c["is_distributor"] else "No",
        ))

    _auto_width(sheet)
    workbook.save(path)


def _auto_width(sheet):
    for col in sheet.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)


def _pdf_table_style():
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a1a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f4")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d6d3d1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ])


def _write_customer_report_pdf(path, report_type, data, period_label=""):
    from datetime import date
    document = SimpleDocTemplate(
        str(path), pagesize=letter,
        topMargin=40, bottomMargin=36, leftMargin=36, rightMargin=36,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    h2_style = styles["Heading2"]
    normal = styles["Normal"]
    normal.fontSize = 8

    TITLES = {
        "customers": "Reporte de Clientes Activos",
        "customer_geo": "Reporte de Distribución Geográfica",
        "international_customers": "Reporte de Clientes Internacionales",
    }

    elements = [
        Paragraph("Juhnios Rold", title_style),
        Paragraph(TITLES.get(report_type, "Reporte de Clientes"), h2_style),
        Paragraph(f"Generado el {date.today().strftime('%d/%m/%Y')}", normal),
    ]
    if period_label:
        elements.append(Paragraph(f"Período exportado: {period_label}", normal))
    elements.append(Spacer(1, 16))

    mode_map = {"WHOLESALE": "Mayorista", "RETAIL": "Retail"}

    if report_type == "customers":
        # Segmentación summary
        segments = data.get("customer_segments", [])
        if segments:
            elements.append(Paragraph("Segmentación de clientes", h2_style))
            elements.append(Spacer(1, 6))
            seg_rows = [["Segmento", "Cantidad", "Porcentaje"]] + [
                [s["segment"], str(s["count"]), f"{s['percentage']}%"]
                for s in segments
            ]
            seg_table = Table(seg_rows, colWidths=[7 * cm, 4 * cm, 4 * cm])
            seg_table.setStyle(_pdf_table_style())
            elements += [seg_table, Spacer(1, 20)]

        customers = data.get("top_customers", [])
        if customers:
            elements.append(Paragraph("Detalle de clientes activos", h2_style))
            elements.append(Spacer(1, 6))
            rows = [["#", "Nombre", "Ciudad", "Pedidos", "Ingresos", "Ticket prom.", "Segmento", "Modo"]]
            for i, c in enumerate(customers, 1):
                rows.append([
                    str(i), c["name"][:22], c["city"] or "—",
                    str(c["orders"]), _fmt_cop(c["revenue"]), _fmt_cop(c["avg_ticket"]),
                    c["segment"], mode_map.get(c["mode"], c["mode"]),
                ])
            col_widths = [1 * cm, 5 * cm, 3 * cm, 1.8 * cm, 3 * cm, 3 * cm, 2.5 * cm, 2.5 * cm]
            table = Table(rows, colWidths=col_widths, repeatRows=1)
            table.setStyle(_pdf_table_style())
            elements.append(table)

    elif report_type == "customer_geo":
        geo = data.get("customer_geo", [])
        if geo:
            elements.append(Paragraph("Top ciudades por ingresos", h2_style))
            elements.append(Spacer(1, 6))
            rows = [["#", "Ciudad", "Clientes", "Pedidos", "Ingresos totales", "% de la base"]]
            for i, g in enumerate(geo, 1):
                rows.append([str(i), g["city"], str(g["customers"]), str(g["orders"]),
                              _fmt_cop(g["revenue"]), f"{g['percentage']}%"])
            col_widths = [1 * cm, 5 * cm, 3 * cm, 3 * cm, 5 * cm, 3 * cm]
            table = Table(rows, colWidths=col_widths, repeatRows=1)
            table.setStyle(_pdf_table_style())
            elements.append(table)

    elif report_type == "international_customers":
        intl = data.get("international_customers", [])
        elements.append(Paragraph(f"Total: {len(intl)} cliente(s) internacionales detectado(s)", normal))
        elements.append(Spacer(1, 10))
        if intl:
            rows = [["#", "Nombre", "País(es)", "Pedidos", "Ingresos", "Modo", "Distribuidor"]]
            for i, c in enumerate(intl, 1):
                rows.append([
                    str(i), c["name"][:24], ", ".join(c["countries"]),
                    str(c["orders"]), _fmt_cop(c["revenue"]),
                    mode_map.get(c["mode"], c["mode"]),
                    "Sí" if c["is_distributor"] else "No",
                ])
            col_widths = [1 * cm, 5 * cm, 4 * cm, 2 * cm, 3.5 * cm, 2.5 * cm, 2.5 * cm]
            table = Table(rows, colWidths=col_widths, repeatRows=1)
            table.setStyle(_pdf_table_style())
            elements.append(table)
        else:
            elements.append(Paragraph("No se encontraron clientes internacionales.", normal))

    document.build(elements)


def _write_xlsx(path, report_type, data):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = report_type[:31]
    sheet.append(("Indicador", "Valor"))
    for key, value in data.items():
        sheet.append((key, json.dumps(value, default=str, ensure_ascii=False) if isinstance(value, (list, dict)) else value))
    workbook.save(path)


def _write_pdf(path, report_type, data):
    document = canvas.Canvas(str(path), pagesize=letter)
    document.setTitle(f"Reporte {report_type}")
    document.drawString(50, 750, f"Juhnios Rold - Reporte {report_type}")
    y = 720
    for key, value in data.items():
        text = f"{key}: {json.dumps(value, default=str, ensure_ascii=False)}"
        document.drawString(50, y, text[:100])
        y -= 22
        if y < 50:
            document.showPage()
            y = 750
    document.save()


@shared_task
def export_sales_report(output_format="xlsx", date_from=None, date_to=None, status=None, client_type=None):
    data = SalesReportQuery().execute(
        date_from=date_from, date_to=date_to, status=status, client_type=client_type,
    )
    report_dir = Path(settings.MEDIA_ROOT) / "reports"
    report_dir.mkdir(parents=True, exist_ok=True)
    filename = f"reporte-ventas-{uuid4().hex}.{output_format}"
    output_path = report_dir / filename

    period_label = _build_period_label({
        "date_from": date_from, "date_to": date_to, "status": status, "client_type": client_type,
    })

    if output_format == "pdf":
        _write_sales_report_pdf(output_path, data, period_label)
    else:
        _write_sales_report_xlsx(output_path, data, period_label)

    return {
        "status": "generated",
        "url": f"{settings.MEDIA_URL}reports/{filename}",
    }


def _write_sales_report_xlsx(path, data, period_label=""):
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Ventas mensuales"
    if period_label:
        sheet.append((f"Período exportado: {period_label}",))
        sheet.append(())
    sheet.append(("Mes", "Ventas", "Pedidos"))
    for row in data["monthly_sales"]:
        sheet.append((_month_label(row["month"]), row["total"], row["orders"]))

    categories_sheet = workbook.create_sheet("Por categoría")
    categories_sheet.append(("Categoría", "Ventas"))
    for row in data["sales_by_category"]:
        categories_sheet.append((row["category"], row["total"]))

    products_sheet = workbook.create_sheet("Top productos")
    products_sheet.append(("Producto", "Unidades", "Ingresos"))
    for row in data["top_products"]:
        products_sheet.append((row["name"], row["units"], row["revenue"]))

    segments_sheet = workbook.create_sheet("Segmentación clientes")
    segments_sheet.append(("Segmento", "Cantidad", "Porcentaje"))
    for row in data["customer_segments"]:
        segments_sheet.append((row["segment"], row["count"], row["percentage"]))

    workbook.save(path)


def _line_chart_drawing(monthly_sales):
    drawing = Drawing(480, 220)
    chart = HorizontalLineChart()
    chart.x = 50
    chart.y = 30
    chart.width = 400
    chart.height = 160
    chart.data = [[float(row["total"]) for row in monthly_sales]]
    chart.categoryAxis.categoryNames = [_month_label(row["month"]) for row in monthly_sales]
    chart.categoryAxis.labels.fontSize = 8
    chart.valueAxis.labels.fontSize = 8
    chart.valueAxis.valueMin = 0
    chart.lines[0].strokeColor = colors.HexColor("#0a0a0a")
    chart.lines[0].strokeWidth = 2
    drawing.add(chart)
    drawing.add(String(50, 200, "Ventas mensuales ($)", fontSize=10))
    return drawing


def _pie_chart_drawing(sales_by_category):
    drawing = Drawing(480, 220)
    chart = Pie()
    chart.x = 80
    chart.y = 20
    chart.width = 150
    chart.height = 150
    chart.data = [float(row["total"]) for row in sales_by_category] or [1]
    categories = [row["category"] for row in sales_by_category] or ["Sin datos"]
    chart.labels = [""] * len(chart.data)
    chart.slices.strokeWidth = 0.5
    for index in range(len(chart.data)):
        chart.slices[index].fillColor = CHART_COLORS[index % len(CHART_COLORS)]
    drawing.add(chart)

    legend = Legend()
    legend.x = 280
    legend.y = 150
    legend.dx = 8
    legend.dy = 8
    legend.fontSize = 8
    legend.alignment = "left"
    legend.colorNamePairs = [
        (CHART_COLORS[index % len(CHART_COLORS)], name) for index, name in enumerate(categories)
    ]
    drawing.add(legend)
    drawing.add(String(80, 195, "Ventas por categoría", fontSize=10))
    return drawing


def _bar_chart_drawing(top_products):
    drawing = Drawing(480, 240)
    chart = VerticalBarChart()
    chart.x = 100
    chart.y = 40
    chart.width = 350
    chart.height = 160
    chart.data = [[float(row["units"]) for row in top_products]]
    chart.categoryAxis.categoryNames = [
        row["name"][:14] for row in top_products
    ]
    chart.categoryAxis.labels.fontSize = 7
    chart.categoryAxis.labels.angle = 30
    chart.categoryAxis.labels.dy = -10
    chart.valueAxis.labels.fontSize = 8
    chart.valueAxis.valueMin = 0
    chart.bars[0].fillColor = colors.HexColor("#0a0a0a")
    drawing.add(chart)
    drawing.add(String(100, 220, "Top 5 productos (unidades)", fontSize=10))
    return drawing


def _write_sales_report_pdf(path, data, period_label=""):
    document = SimpleDocTemplate(str(path), pagesize=letter, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Juhnios Rold - Reporte de ventas", styles["Title"]),
    ]
    if period_label:
        elements.append(Paragraph(f"Período: {period_label}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    if data["monthly_sales"]:
        elements.append(_line_chart_drawing(data["monthly_sales"]))
        elements.append(Spacer(1, 16))

    if data["sales_by_category"]:
        elements.append(_pie_chart_drawing(data["sales_by_category"]))
        elements.append(Spacer(1, 16))

    if data["top_products"]:
        elements.append(_bar_chart_drawing(data["top_products"]))
        elements.append(Spacer(1, 16))

    elements.append(PageBreak())
    elements.append(Paragraph("Segmentación de clientes", styles["Heading2"]))
    elements.append(Spacer(1, 8))
    segment_rows = [["Segmento", "Cantidad", "Porcentaje"]] + [
        [row["segment"], str(row["count"]), f"{row['percentage']}%"]
        for row in data["customer_segments"]
    ]
    segment_table = Table(segment_rows, colWidths=[6 * cm, 4 * cm, 4 * cm])
    segment_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#222222")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
    ]))
    elements.append(segment_table)

    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Top productos", styles["Heading2"]))
    elements.append(Spacer(1, 8))
    product_rows = [["Producto", "Unidades", "Ingresos"]] + [
        [row["name"], str(row["units"]), str(row["revenue"])]
        for row in data["top_products"]
    ]
    product_table = Table(product_rows, colWidths=[8 * cm, 3 * cm, 3 * cm])
    product_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#222222")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
    ]))
    elements.append(product_table)

    document.build(elements)


def _write_inventory_report_xlsx(path, report_type, title, data, period_label=""):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title[:31]
    header_row = _write_period_label(sheet, period_label)

    if report_type == "compras":
        headers = ["N° OC", "Proveedor", "Estado", "Emitida", "Esperada", "Bodega destino",
                   "Total", "Cant. pedida", "Cant. recibida", "Cant. pendiente"]
        _apply_header(sheet, headers, start_row=header_row)
        for i, o in enumerate(data["orders"], 1):
            _apply_row(sheet, header_row + i, (
                o["number"], o["supplier"], o["status_label"], o["issued_at"], o["expected_at"] or "—",
                o["destination"], _fmt_cop(o["total"]), o["ordered_quantity"], o["received_quantity"], o["pending_quantity"],
            ))
    elif report_type == "bajo-minimo":
        headers = ["SKU", "Producto", "Presentación", "Ubicación", "Existencia", "Mínimo", "Faltante"]
        _apply_header(sheet, headers, start_row=header_row)
        for i, item in enumerate(data["items"], 1):
            _apply_row(sheet, header_row + i, (
                item["sku"], item["product"], item["presentation"], item["location"],
                item["quantity"], item["minimum_quantity"], item["shortage"],
            ))
    elif report_type == "produccion":
        headers = ["N° OP", "Fórmula", "Producto", "Estado", "Inicio", "Cierre",
                   "Planeado", "Real", "Variación", "% Rendimiento"]
        _apply_header(sheet, headers, start_row=header_row)
        for i, o in enumerate(data["orders"], 1):
            _apply_row(sheet, header_row + i, (
                o["number"], o["formula"], o["output_item"], o["status_label"],
                o["started_at"] or "—", o["closed_at"] or "—",
                o["planned_quantity"], o["actual_quantity"], o["variance"], f"{o['yield_percentage']}%",
            ))
    elif report_type == "mermas":
        headers = ["Fecha", "Producto", "Ubicación", "Tipo", "Cantidad", "Motivo"]
        _apply_header(sheet, headers, start_row=header_row)
        for i, m in enumerate(data["movements"], 1):
            _apply_row(sheet, header_row + i, (
                m["date"], m["product"], m["location"], m["type_label"], m["quantity"], m["reason"],
            ))
    elif report_type == "corte":
        headers = ["SKU", "Producto", "Ubicación", "Existencia actual", f"Existencia al {data['cutoff_date']}"]
        _apply_header(sheet, headers, start_row=header_row)
        for i, item in enumerate(data["items"], 1):
            _apply_row(sheet, header_row + i, (
                item["sku"], item["product"], item["location"], item["current_quantity"], item["quantity_at_date"],
            ))
    elif report_type in ("inv-general", "inv-bodega", "inv-grupo", "valorizado"):
        headers = ["SKU", "Producto", "Categoría", "Ubicación", "Existencia", "Costo unit.",
                   "Valor sin IVA", "Valor con IVA", "Precio venta", "Valor a precio venta"]
        _apply_header(sheet, headers, start_row=header_row)
        for i, item in enumerate(data["items"], 1):
            _apply_row(sheet, header_row + i, (
                item["sku"], item["product"], item["category"], item["location"], item["quantity"],
                _fmt_cop(item["unit_cost"]), _fmt_cop(item["value_no_vat"]), _fmt_cop(item["value_with_vat"]),
                _fmt_cop(item["sale_price"]), _fmt_cop(item["value_at_sale_price"]),
            ))
        summary_row = header_row + len(data["items"]) + 2
        sheet.cell(row=summary_row, column=1, value="Totales:")
        sheet.cell(row=summary_row, column=7, value=_fmt_cop(data["summary"]["total_value_no_vat"]))
        sheet.cell(row=summary_row, column=8, value=_fmt_cop(data["summary"]["total_value_with_vat"]))
        if "total_value_at_sale_price" in data["summary"]:
            sheet.cell(row=summary_row, column=10, value=_fmt_cop(data["summary"]["total_value_at_sale_price"]))
    else:  # movimientos
        headers = ["Fecha", "Producto", "Ubicación", "Tipo", "Cantidad", "Motivo", "Registrado por"]
        _apply_header(sheet, headers, start_row=header_row)
        for i, m in enumerate(data["movements"], 1):
            _apply_row(sheet, header_row + i, (
                m["date"], m["product"], m["location"], m["type_label"], m["quantity"], m["reason"], m["created_by"],
            ))

    _auto_width(sheet)
    workbook.save(path)


def _write_inventory_report_pdf(path, report_type, title, data, period_label=""):
    from datetime import date
    document = SimpleDocTemplate(
        str(path), pagesize=letter,
        topMargin=40, bottomMargin=36, leftMargin=36, rightMargin=36,
    )
    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    normal.fontSize = 8

    elements = [
        Paragraph("Juhnios Rold", styles["Title"]),
        Paragraph(title, styles["Heading2"]),
        Paragraph(f"Generado el {date.today().strftime('%d/%m/%Y')}", normal),
    ]
    if period_label:
        elements.append(Paragraph(f"Filtros: {period_label}", normal))
    elements.append(Spacer(1, 16))

    if report_type == "compras":
        rows = [["N° OC", "Proveedor", "Estado", "Emitida", "Bodega", "Total", "Pendiente"]] + [
            [o["number"], o["supplier"][:22], o["status_label"], o["issued_at"], o["destination"][:16],
             _fmt_cop(o["total"]), str(o["pending_quantity"])]
            for o in data["orders"]
        ]
        col_widths = [2.2 * cm, 4 * cm, 2 * cm, 2.2 * cm, 3 * cm, 3 * cm, 2.5 * cm]
    elif report_type == "bajo-minimo":
        rows = [["SKU", "Producto", "Ubicación", "Existencia", "Mínimo", "Faltante"]] + [
            [i["sku"], i["product"][:26], i["location"][:16], str(i["quantity"]), str(i["minimum_quantity"]), str(i["shortage"])]
            for i in data["items"]
        ]
        col_widths = [2.2 * cm, 5.5 * cm, 3.5 * cm, 2.3 * cm, 2 * cm, 2.3 * cm]
    elif report_type == "produccion":
        rows = [["N° OP", "Producto", "Estado", "Cierre", "Planeado", "Real", "% Rend."]] + [
            [o["number"], o["output_item"][:22], o["status_label"], o["closed_at"] or "—",
             str(o["planned_quantity"]), str(o["actual_quantity"]), f"{o['yield_percentage']}%"]
            for o in data["orders"]
        ]
        col_widths = [2.2 * cm, 4.5 * cm, 2.3 * cm, 2.2 * cm, 2.3 * cm, 2.3 * cm, 2.2 * cm]
    elif report_type == "mermas":
        rows = [["Fecha", "Producto", "Ubicación", "Tipo", "Cantidad"]] + [
            [m["date"], m["product"][:24], m["location"][:16], m["type_label"], str(m["quantity"])]
            for m in data["movements"]
        ]
        col_widths = [2.3 * cm, 5 * cm, 3.5 * cm, 3 * cm, 2.5 * cm]
    elif report_type == "corte":
        rows = [["SKU", "Producto", "Ubicación", "Actual", f"Al {data['cutoff_date']}"]] + [
            [i["sku"], i["product"][:26], i["location"][:16], str(i["current_quantity"]), str(i["quantity_at_date"])]
            for i in data["items"]
        ]
        col_widths = [2.2 * cm, 5.5 * cm, 3.5 * cm, 2.5 * cm, 2.5 * cm]
    elif report_type in ("inv-general", "inv-bodega", "inv-grupo", "valorizado"):
        rows = [["SKU", "Producto", "Categoría", "Ubicación", "Exist.", "Costo unit.", "Valor sin IVA", "Precio venta"]] + [
            [i["sku"], i["product"][:20], i["category"][:14], i["location"][:14], str(i["quantity"]),
             _fmt_cop(i["unit_cost"]), _fmt_cop(i["value_no_vat"]), _fmt_cop(i["sale_price"])]
            for i in data["items"]
        ]
        col_widths = [1.8 * cm, 3.6 * cm, 2.6 * cm, 2.6 * cm, 1.6 * cm, 2.4 * cm, 2.6 * cm, 2.4 * cm]
    else:  # movimientos
        rows = [["Fecha", "Producto", "Ubicación", "Tipo", "Cantidad", "Motivo"]] + [
            [m["date"], m["product"][:22], m["location"][:16], m["type_label"], str(m["quantity"]), m["reason"][:20]]
            for m in data["movements"]
        ]
        col_widths = [2.2 * cm, 4.5 * cm, 3.2 * cm, 2.8 * cm, 2.2 * cm, 3.6 * cm]

    if len(rows) > 1:
        table = Table(rows, colWidths=col_widths, repeatRows=1)
        table.setStyle(_pdf_table_style())
        elements.append(table)
    else:
        elements.append(Paragraph("Sin datos para los filtros seleccionados.", normal))

    document.build(elements)
