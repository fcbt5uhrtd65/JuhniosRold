import re
import unicodedata
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from openpyxl import load_workbook

from apps.employees.infrastructure.models import Employee
from apps.human_resources.application.use_cases import CalculatePayrollPeriod
from apps.human_resources.infrastructure.models import (
    Attendance,
    BiometricDevice,
    EmployeeBiometricId,
    PayrollPeriod,
)


SKIPPED_SHEET_WORDS = ("MARCACIONES", "TIMBRADAS")
SKIPPED_NAME_TOKENS = {"FINCA", "CENTRO", "REVISION"}
TOTALS_LABEL = "TOTALES"


def normalize(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Z0-9 ]+", " ", text.upper())
    return re.sub(r"\s+", " ", text).strip()


def decimal_value(value, default="0"):
    if value in (None, ""):
        return Decimal(default)
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError):
        return Decimal(default)


def as_datetime(value):
    if isinstance(value, datetime):
        return value
    return None


def aware(value):
    if value is None or timezone.is_aware(value):
        return value
    return timezone.make_aware(value)


def employee_full_name(employee):
    return f"{employee.first_name} {employee.last_name}".strip() or employee.employee_code


class WorkbookEmployeeSheet:
    def __init__(self, worksheet):
        self.worksheet = worksheet
        self.title = worksheet.title
        self.codes = set(re.findall(r"\d+", self.title))
        self.name_hint = self._name_hint()
        self.rows = []
        self.totals = {}

    def _name_hint(self):
        title_without_codes = re.sub(r"\d+", " ", self.title)
        tokens = [token for token in normalize(title_without_codes).split() if token not in SKIPPED_NAME_TOKENS]
        return " ".join(tokens)

    def parse(self):
        last_code = None
        for row in range(4, self.worksheet.max_row + 1):
            first = self.worksheet.cell(row=row, column=1).value
            first_label = normalize(first)
            if first_label == TOTALS_LABEL:
                self.totals = self._totals(row)
                break

            entrada = as_datetime(self.worksheet.cell(row=row, column=2).value)
            salida = as_datetime(self.worksheet.cell(row=row, column=3).value)
            if first not in (None, ""):
                code_match = re.search(r"\d+", str(first))
                if code_match:
                    last_code = code_match.group(0)
                    self.codes.add(last_code)
            code = last_code or (sorted(self.codes)[0] if self.codes else "")

            if not entrada and not salida:
                continue

            descanso_hours = decimal_value(self.worksheet.cell(row=row, column=9).value)
            break_start = as_datetime(self.worksheet.cell(row=row, column=7).value)
            break_end = as_datetime(self.worksheet.cell(row=row, column=8).value)
            if entrada and salida and salida < entrada:
                salida += timedelta(days=1)
            if break_start and break_end and break_end < break_start:
                break_end += timedelta(days=1)

            if descanso_hours > 0 and entrada and salida:
                descanso_delta = timedelta(hours=float(descanso_hours))
                if break_start:
                    break_end = break_start + descanso_delta
                elif break_end and not break_start:
                    break_start = break_end - descanso_delta
                elif not break_start and not break_end:
                    worked_without_break = max(salida - entrada - descanso_delta, timedelta())
                    break_start = entrada + (worked_without_break / 2)
                    break_end = break_start + descanso_delta

            self.rows.append({
                "row": row,
                "code": code,
                "date": (entrada or salida).date(),
                "check_in": aware(entrada),
                "check_out": aware(salida),
                "break_start": aware(break_start),
                "break_end": aware(break_end),
                "worked_hours": decimal_value(self.worksheet.cell(row=row, column=5).value),
                "ordinary_hours": decimal_value(self.worksheet.cell(row=row, column=6).value),
                "break_hours": descanso_hours,
                "extra_day_hours": decimal_value(self.worksheet.cell(row=row, column=10).value),
                "extra_night_hours": decimal_value(self.worksheet.cell(row=row, column=11).value),
                "night_surcharge_hours": decimal_value(self.worksheet.cell(row=row, column=12).value),
                "sunday_day_hours": decimal_value(self.worksheet.cell(row=row, column=13).value),
                "sunday_extra_day_hours": decimal_value(self.worksheet.cell(row=row, column=14).value),
                "sunday_night_hours": decimal_value(self.worksheet.cell(row=row, column=15).value),
                "sunday_extra_night_hours": decimal_value(self.worksheet.cell(row=row, column=16).value),
                "incapacity_days": decimal_value(self.worksheet.cell(row=row, column=17).value),
            })

        return self

    def _totals(self, row):
        return {
            "worked_hours": decimal_value(self.worksheet.cell(row=row, column=5).value),
            "ordinary_hours": decimal_value(self.worksheet.cell(row=row, column=6).value),
            "break_hours": decimal_value(self.worksheet.cell(row=row, column=9).value),
            "extra_day_hours": decimal_value(self.worksheet.cell(row=row, column=10).value),
            "extra_night_hours": decimal_value(self.worksheet.cell(row=row, column=11).value),
            "night_surcharge_hours": decimal_value(self.worksheet.cell(row=row, column=12).value),
            "sunday_day_hours": decimal_value(self.worksheet.cell(row=row, column=13).value),
            "sunday_extra_day_hours": decimal_value(self.worksheet.cell(row=row, column=14).value),
            "sunday_night_hours": decimal_value(self.worksheet.cell(row=row, column=15).value),
            "sunday_extra_night_hours": decimal_value(self.worksheet.cell(row=row, column=16).value),
            "incapacity_days": decimal_value(self.worksheet.cell(row=row, column=17).value),
        }


class Command(BaseCommand):
    help = "Importa el Excel legado de nomina por hojas de empleado y crea mapeos/asistencias."

    def add_arguments(self, parser):
        parser.add_argument("workbook_path")
        parser.add_argument("--period-start", required=True, help="Fecha inicial del periodo, YYYY-MM-DD.")
        parser.add_argument("--period-end", required=True, help="Fecha final del periodo, YYYY-MM-DD.")
        parser.add_argument("--device-name", default="Huellero legado Excel")
        parser.add_argument("--commit", action="store_true", help="Escribe cambios. Sin esto solo muestra vista previa.")
        parser.add_argument("--overwrite", action="store_true", help="Actualiza asistencias existentes no corregidas manualmente.")
        parser.add_argument("--overwrite-corrected", action="store_true", help="Tambien sobrescribe asistencias corregidas manualmente.")
        parser.add_argument("--create-period", action="store_true", help="Crea o reutiliza PayrollPeriod del rango.")
        parser.add_argument("--calculate-payroll", action="store_true", help="Calcula nomina para empleados importados.")

    def handle(self, *args, **options):
        workbook_path = Path(options["workbook_path"])
        if not workbook_path.exists():
            raise CommandError(f"No existe el archivo: {workbook_path}")

        try:
            period_start = datetime.strptime(options["period_start"], "%Y-%m-%d").date()
            period_end = datetime.strptime(options["period_end"], "%Y-%m-%d").date()
        except ValueError as exc:
            raise CommandError("Las fechas deben venir en formato YYYY-MM-DD.") from exc
        if period_end < period_start:
            raise CommandError("period-end no puede ser menor que period-start.")

        workbook = load_workbook(workbook_path, data_only=True, read_only=False)
        sheets = [
            WorkbookEmployeeSheet(ws).parse()
            for ws in workbook.worksheets
            if ws.sheet_state == "visible"
            and not any(word in normalize(ws.title) for word in SKIPPED_SHEET_WORDS)
        ]
        sheets = [sheet for sheet in sheets if sheet.rows and sheet.codes]

        device = None
        if options["commit"]:
            device, _ = BiometricDevice.objects.get_or_create(
                name=options["device_name"],
                defaults={"location": "Importado desde Excel legado de nomina"},
            )

        summary = {
            "sheets": len(sheets),
            "matched": 0,
            "unmatched": 0,
            "mappings_created": 0,
            "attendance_created": 0,
            "attendance_updated": 0,
            "attendance_skipped": 0,
        }
        imported_employee_ids = set()
        warnings = []

        with transaction.atomic():
            for sheet in sheets:
                employee = self._resolve_employee(sheet, device)
                if not employee:
                    summary["unmatched"] += 1
                    warnings.append(f"Sin empleado unico para hoja '{sheet.title}' codigos {', '.join(sorted(sheet.codes))}")
                    continue

                summary["matched"] += 1
                imported_employee_ids.add(employee.id)
                mapping_result = self._ensure_mappings(sheet, employee, device, options["commit"])
                summary["mappings_created"] += mapping_result["created"]
                warnings.extend(mapping_result["warnings"])

                attendance_result = self._import_attendance(sheet, employee, period_start, period_end, options)
                for key in ("attendance_created", "attendance_updated", "attendance_skipped"):
                    summary[key] += attendance_result[key]

                self.stdout.write(
                    f"{sheet.title}: {employee_full_name(employee)} | filas {len(sheet.rows)} | "
                    f"extra diurna Excel {sheet.totals.get('extra_day_hours', 0)}"
                )

            period = None
            if options["create_period"] or options["calculate_payroll"]:
                if not options["commit"]:
                    self.stdout.write("DRY-RUN: se crearia/reutilizaria PayrollPeriod y se calcularia si fue solicitado.")
                else:
                    period, _ = PayrollPeriod.objects.get_or_create(
                        period_start=period_start,
                        period_end=period_end,
                        defaults={"label": f"Nomina {period_start:%d/%m/%Y} - {period_end:%d/%m/%Y}"},
                    )

            if options["calculate_payroll"] and options["commit"] and period:
                employees = Employee.objects.filter(id__in=imported_employee_ids)
                result = CalculatePayrollPeriod().execute(period=period, employee_queryset=employees)
                self.stdout.write(f"Nomina calculada: {result['calculated']} empleado(s), errores: {len(result['errors'])}")
                for error in result["errors"]:
                    warnings.append(f"Error calculando {error['employee']}: {error['error']}")

            if not options["commit"]:
                transaction.set_rollback(True)

        mode = "APLICADO" if options["commit"] else "VISTA PREVIA"
        self.stdout.write(self.style.SUCCESS(f"{mode}: {summary}"))
        for warning in warnings[:80]:
            self.stdout.write(self.style.WARNING(warning))
        if len(warnings) > 80:
            self.stdout.write(self.style.WARNING(f"... {len(warnings) - 80} advertencias mas."))

    def _resolve_employee(self, sheet, device):
        mapped = Employee.objects.filter(
            biometric_ids__biometric_code__in=sheet.codes,
            biometric_ids__is_active=True,
            biometric_ids__deleted_at__isnull=True,
        )
        if device:
            mapped = mapped.filter(Q(biometric_ids__device=device) | Q(biometric_ids__device__isnull=True))
        mapped = mapped.distinct()
        if mapped.count() == 1:
            return mapped.first()

        employee_code_matches = Employee.objects.filter(employee_code__in=sheet.codes)
        if employee_code_matches.count() == 1:
            return employee_code_matches.first()

        if not sheet.name_hint:
            return None
        tokens = sheet.name_hint.split()
        candidates = []
        for employee in Employee.objects.all():
            normalized_name = normalize(employee_full_name(employee))
            if tokens and all(token in normalized_name for token in tokens):
                candidates.append(employee)
        return candidates[0] if len(candidates) == 1 else None

    def _ensure_mappings(self, sheet, employee, device, commit):
        result = {"created": 0, "warnings": []}
        for code in sorted(sheet.codes):
            existing = EmployeeBiometricId.objects.filter(
                device=device,
                biometric_code=code,
                is_active=True,
                deleted_at__isnull=True,
            ).first()
            if existing:
                if existing.employee_id != employee.id:
                    result["warnings"].append(
                        f"Codigo {code} ya apunta a {employee_full_name(existing.employee)}, no a {employee_full_name(employee)}."
                    )
                continue
            if commit:
                EmployeeBiometricId.objects.create(
                    employee=employee,
                    device=device,
                    biometric_code=code,
                    valid_from=timezone.localdate(),
                )
            result["created"] += 1
        return result

    def _import_attendance(self, sheet, employee, period_start, period_end, options):
        result = {"attendance_created": 0, "attendance_updated": 0, "attendance_skipped": 0}
        for row in sheet.rows:
            if row["date"] < period_start or row["date"] > period_end:
                continue
            existing = Attendance.objects.filter(employee=employee, date=row["date"]).first()
            if existing and existing.is_manually_corrected and not options["overwrite_corrected"]:
                result["attendance_skipped"] += 1
                continue
            if existing and not options["overwrite"] and not options["overwrite_corrected"]:
                result["attendance_skipped"] += 1
                continue

            defaults = {
                "check_in": row["check_in"],
                "check_out": row["check_out"],
                "break_start": row["break_start"],
                "break_end": row["break_end"],
                "source": Attendance.Source.BIOMETRIC,
                "has_incomplete_marks": not bool(row["check_in"] and row["check_out"]),
                "notes": (
                    f"Importado desde Excel legado '{sheet.title}' fila {row['row']}. "
                    f"Horas: trabajadas={row['worked_hours']}, ordinarias={row['ordinary_hours']}, "
                    f"descanso={row['break_hours']}, extra_diurna={row['extra_day_hours']}."
                ),
            }
            if options["commit"]:
                _, created = Attendance.objects.update_or_create(
                    employee=employee,
                    date=row["date"],
                    defaults=defaults,
                )
            else:
                created = existing is None
            result["attendance_created" if created else "attendance_updated"] += 1
        return result
